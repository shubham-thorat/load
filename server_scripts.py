import numpy as np
import pandas as pd
from openpyxl import load_workbook
import json
from os import path
import re
import glob


def readFile(filename):
    latencies = []
    try:
        print(f"Reading from file {filename}...")
        with open(file=filename, mode="r") as file:
            for line in file:
                values = line.split()
                latencies.append(int(values[1]))

        print(f"Reading end from file {filename}")
    except FileNotFoundError:
        print(f"File Not Found {filename}")
    except Exception as e:
        print(f"An error occurred readFile: {e}")

    return latencies


def calculate(latencies, file, clients, requests):
    if not latencies:
        print("** Array is empty **")
        return None

    latencies.sort()
    per_99 = np.percentile(latencies, 99)
    per_90 = np.percentile(latencies, 90)
    per_50 = np.percentile(latencies, 50)
    min_value = latencies[0]
    max_value = latencies[-1]

    data = {
        'processed_log_file': file,
        'clients': clients,
        'request_count': len(latencies),
        'min_time': f"{min_value}ms",
        'max_time': f"{max_value}ms",
        "50th_percentile": f"{per_50}ms",
        "90th_percentile": f"{per_90}ms",
        "99th_percentile": f"{per_99}ms",
        # "99.9th percentile": f"{round(per_99_9 / 1000, 2)}ms",
    }
    print(f'Latencies data for log file {file} : \n{data}')
    return data


def addToJSONFile(filePath, latenciesData):
    try:
        if path.isfile(filePath) is False:
            fp = open(filePath, "w")
            fp.close()

        dictObj = []
        with open(filePath) as fp:
            try:
                dictObj = json.load(fp)
            except json.JSONDecodeError:
                pass

        dictObj.append(latenciesData)

        with open(filePath, 'w') as json_file:
            json.dump(dictObj, json_file,
                      indent=2,
                      separators=(',', ': '))

        print('Successfully written to the JSON file\n\n')

    except Exception as e:
        print(f"An error occurred addToJSONFile: {e}")


def convertJSONToArray(item):
    # 'processed_log_file': file,
    #     'clients': clients,
    #     'request_count': len(latencies),
    #     'min_time': f"{round(min_value, 4)}ms",
    #     'max_time': f"{round(max_value, 4)}ms",
    #     "50th_percentile": f"{round(per_50, 4)}ms",
    #     "90th_percentile": f"{round(per_90, 4)}ms",
    #     "99th_percentile": f"{round(per_99, 4)}ms",
    try:
        return [
            item['request_count'],
            item['clients'],
            item['min_time'],
            item['max_time'],
            item["50th_percentile"],
            item["90th_percentile"],
            item["99th_percentile"],
        ]
    except Exception as e:
        print('Error while parsing object', e)


def addJSONToExcel(inputJSONFile, outputXlsxFile):
    try:
        data = {}
        with open(inputJSONFile) as json_file:
            data = json.load(json_file)

        wb = load_workbook(outputXlsxFile)

        sheet_name = 'Sheet2'  # Change this to the desired sheet name
        sheet = wb[sheet_name]
        index = sheet.max_row + 1
        for item in data:
            try:
                row = convertJSONToArray(item)
                row.insert(0, index)
                sheet.append(row)
                index += 1
            except Exception as e:
                print(f"Error while appending row to excel {item} {e}")

        wb.save(outputXlsxFile)
    except Exception as e:
        print(f"An error occurred addJSONToExcel: {e}")


def extractNumber(filePath):
    pattern = r'\d+'

    numbers = re.findall(pattern, filePath)
    if numbers:
        return [int(numbers[0]), int(numbers[1])]
    else:
        return None


def main2():
    files = input('Enter space separated file path to process : ')
    outputFile = input('Enter Ouput JSON File : ') or "output.json"
    fileNames = re.split(' +', files)

    for file in fileNames:
        latencies = readFile(filename=file)
        latenciesData = calculate(latencies, file)
        if latenciesData:
            addToJSONFile(outputFile, latenciesData)


def main():
    try:
        # Find all files matching the pattern "output_*.log"
        file_pattern = "./output/logs/output_server_*.json"
        matching_files = glob.glob(file_pattern)
        output_file = './output/json/result_server.json'
        # Read the content of each matching file
        print(matching_files)
        if not matching_files:
            print("NO matching file founds")
            return

        for file_path in matching_files:
            [clients, requests] = extractNumber(filePath=file_path)
            latencies = readFile(filename=file_path)
            result = calculate(latencies, file=file_path,
                               clients=clients, requests=requests)
            addToJSONFile(output_file, result)

        print(
            f'** Processed logs file & result added to JSON File {output_file}**')
        excel_file = 'load_test_grpc.xlsx'
        print("hello")
        addJSONToExcel(inputJSONFile=output_file, outputXlsxFile=excel_file)
        print(f'Succesfully added to Excel file {excel_file}')
    except Exception as e:
        print(f"An error occurred main method : {e}")


main()
